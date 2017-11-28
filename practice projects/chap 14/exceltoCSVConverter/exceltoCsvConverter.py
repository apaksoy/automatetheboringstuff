#!/usr/bin/env python3
'''Excel can save a spreadsheet to a CSV file with a few mouse clicks,
but if you had to convert hundreds of Excel files to CSVs, it would take
hours of clicking. Using the openpyxl module from Chapter  12, write a
program that reads all the Excel files in the current working directory
and outputs them as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create
one CSV file per sheet. The filenames of the CSV files should be <excel
filename>_<sheet title>.csv, where <excel filename> is the filename of
the Excel file without the file extension (for example, 'spam_data', not
'spam_data.xlsx') and <sheet title> is the string from the Worksheet
object’s title variable.
'''

import csv
import os

import openpyxl


# Skip non-xlsx files, load the workbook object.
for filename in os.listdir('.'):
    if not filename.endswith('.xlsx'):
        continue  # skip non-xlsx files
    
    wb = openpyxl.load_workbook(filename, read_only = True)
    
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename
        # and sheet title.
        csvFilename = filename.rstrip('.xlsx') +\
                      '_' + sheetName + '.csv'
        csvFileObj = open(csvFilename, 'w')
        
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData = rowData + [sheet.cell\
                          (row=rowNum, column=colNum).value]               
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFileObj.close()

