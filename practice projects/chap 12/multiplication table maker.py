#!/usr/bin/env python3
'''
Creates an N by N multiplication table in excel where N is an integer
specified by the user. Saves the table when finished.
usage: python3 'multiplication table maker.py' <number less than N>
'''
'''
Create a program multiplicationTable.py that takes a number N from the
command line and creates an N×N multiplication table in an Excel
spreadsheet. For example, when the program is run like this:

py multiplicationTable.py 6
... it should create a spreadsheet that looks like
'''

import openpyxl
import sys

def usage (number):
    return print(f'usage: python3 "multiplication table maker.py" <number less than {number}>')
    
maxNumber = 100
filename = 'multiplication table.xlsx'

if len(sys.argv) < 2:
    usage(maxNumber)
    sys.exit()

try:
    tableNumber = int(sys.argv[1])
    if tableNumber >= maxNumber:
        print(f'You entered "{sys.argv[1]}" as the number')
        print(f'Pls enter a number less than {maxNumber}')
        usage(maxNumber)
        sys.exit()
except ValueError:
    print(f'You entered "{sys.argv[1]}" as a number,'\
          f' pls enter a number')
    usage(maxNumber)
    sys.exit()

print(f'creating {tableNumber} by {tableNumber} multiplication table') 
wb = openpyxl.Workbook()
sheet = wb.active

# fill row and column headers
for i in range(1, tableNumber + 1):
    sheet.cell(row = 1, column = i + 1).value = i
    sheet.cell(row = i + 1, column = 1).value = i

# fill the table
for i in range(2, tableNumber + 2):
    for j in range(2, tableNumber + 2):
        sheet.cell(row = i , column = j).value = (i - 1) * (j - 1)

wb.save(filename)
print(f'multiplication table saved as "{filename}"')
