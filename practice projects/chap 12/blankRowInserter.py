#!/usr/bin/env python3
'''
Create a program blankRowInserter.py that takes two integers and a
filename string as command line arguments. Let’s call the first integer
N and the second integer M. Starting at row N, the program should insert
M blank rows into the spreadsheet. For example, when the program is run
like this:

python blankRowInserter.py 3 2 myProduce.xlsx

... the “before” and “after” spreadsheets should look like Figure 12-12.
'''

import openpyxl
import re
import sys

def usage ():
    return print('usage: python3 "blank line inserter.py" '\
                 '<insertion_line_number> <number_of_lines> <file_name>')

# check if enough number of arguments provided at the command line
if len(sys.argv) < 4:
    usage()
    sys.exit()

# pass arguments to relevant variables
inNumber = int(sys.argv[1])
noInsert = int(sys.argv[2])
fname = sys.argv[3]

# check if provided filename to open makes sense
try:
    mo = re.search(r'([\w_-]+)\.(xlsx)$', fname, re.I)
    firstName, lastName = mo.group(1), mo.group(2)
except Exception:
    print(f'"{sys.argv[3]}" is not a proper excel file name for openpyxl')
    sys.exit()
    
wb = openpyxl.load_workbook(filename=fname)
sheet = wb.active

# check if insertion line number makes sense 
if sheet.max_row <= inNumber:
    print('No insertion necessary')
    print(f'Insertion line number {inNumber} greater than existing '\
          f'number of lines of {sheet.max_row} in file {fname}')
    usage()
    sys.exit()

# move cells to new places starting with the last row
for i in range(sheet.max_row, inNumber - 1, -1):
    for j in range(1, sheet.max_column + 1):
        sheet.cell(row = i + noInsert, column = j).value\
            = sheet.cell(row=i, column=j).value
        sheet.cell(row=i, column=j).value = '' # this really make those
                                               # cells empty

upFilename = firstName + '_updated.' + lastName        
wb.save(upFilename)
print(f'Modified file saved under {upFilename}') 
    
