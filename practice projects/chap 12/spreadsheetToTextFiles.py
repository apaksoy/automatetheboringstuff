#!/usr/bin/env python3
'''
Write a program that performs the tasks of the previous program in
reverse order: The program should open a spreadsheet and write the
cells of column A into one text file, the cells of column B into
another text file, and so on.
'''

import openpyxl

wb = openpyxl.load_workbook('SpreadsheettoText.xlsx')
sheet = wb.active

for i in range(1, sheet.max_column + 1):
    txtfileObj = open(f'TexttoSpreadsheetOut{i}.txt', 'w', encoding='utf-8')
    for j in range(1, sheet.max_row + 1):
        c = sheet.cell(row=j, column=i).value
        if c:
            txtfileObj.write(str(c))
    txtfileObj.close()


### just another solution to empty cells issue
### returns index for the last non-empty cell 
### in a Worksheet object created via openpyxl module
##def get_size(wObj):
##    for counter, cell in enumerate(wObj):
##        if cell.value:
##            bottom = counter
##    return bottom + 1

##colsObj = sheet.columns
##for i, item in enumerate(colsObj):
##    txtfileObj = open(f'TexttoSpreadsheetOut{i + 1}.txt', 'w', encoding='utf-8')
##    colSize = get_size(item)
##    for c in item[0:colSize]:
##        if c.value:
##            txtfileObj.write(str(c.value))
##    txtfileObj.close()
