#!/usr/bin/env python3
'''
Write a program to invert the row and column of the cells in the
spreadsheet. For example, the value at row 5, column 3 will be at row 3,
column 5 (and vice versa). This should be done for all cells in the
spreadsheet.
'''

import openpyxl

fname = "toBeCellInverted.xlsx"
wb = openpyxl.load_workbook(fname)
sheet = wb.active

# get size of table to be inverted
rSize = sheet.max_row
cSize = sheet.max_column

# generate list to store table temporarily
# tableList = [[''] * cSize for i in range(rSize)]
# found a way to get a 2-d list without having to declare its size
# beforehand
tableList = []

# store table to list and empty the cells
for i in range(1, rSize + 1):
    tableList.append([])
    for j in range(1, cSize + 1):
        tableList[i - 1].append(sheet.cell(row=i, column=j).value)
        sheet.cell(row=i, column=j).value = ''
        
# write table to cells as transposed         
for i in range(1, rSize + 1):
    for j in range(1, cSize + 1):
        sheet.cell(row=j, column=i).value = tableList[i - 1][j - 1]

wb.save(fname)
